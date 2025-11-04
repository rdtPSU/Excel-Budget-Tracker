from flask import Flask, render_template, request, redirect, url_for
from openpyxl import load_workbook
from datetime import datetime

app = Flask(__name__)

@app.route("/", methods=["GET", "POST"])
def home():
    if request.method == "POST":
        category = request.form["category"]
        amount = request.form["amount"]
        date = datetime.now().strftime("%Y-%m-%d")

        workbook = load_workbook("budget.xlsx")
        sheet = workbook.active

        next_row = sheet.max_row + 1
        sheet[f"A{next_row}"] = date
        sheet[f"B{next_row}"] = category
        sheet[f"C{next_row}"] = float(amount)  # ensure it saves as a number

        workbook.save("budget.xlsx")

        # Redirect back to home after submission
        return redirect(url_for("success"))

    return render_template("index.html")

@app.route("/success")
def success():
    # This page just shows confirmation briefly then redirects
    return render_template("success.html")

if __name__ == "__main__":
    app.run(debug=True)
